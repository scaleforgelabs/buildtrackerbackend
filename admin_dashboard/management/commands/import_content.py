"""
Management command to import content ideas from the Excel file.
Usage: python manage.py import_content /path/to/BuildTracker_Content_Idea_Bank_v2.xlsx
"""
import openpyxl
from django.core.management.base import BaseCommand
from admin_dashboard.models import ContentPost


PLATFORM_MAP = {
    'Twitter/X': 'twitter',
    'Instagram Post': 'instagram_post',
    'Instagram Slides': 'instagram_slides',
    'LinkedIn': 'linkedin',
}

STATUS_MAP = {
    'To Do': 'to_do',
    'In Progress': 'in_progress',
    'Scheduled': 'scheduled',
    'Posted': 'posted',
    'Skipped': 'skipped',
}

PRIORITY_MAP = {
    'Low': 'low',
    'Medium': 'medium',
    'High': 'high',
}


class Command(BaseCommand):
    help = 'Import content ideas from BuildTracker_Content_Idea_Bank_v2.xlsx'

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help='Path to the Excel file')
        parser.add_argument('--clear', action='store_true', help='Clear existing records before import')

    def handle(self, *args, **options):
        filepath = options['filepath']
        if options['clear']:
            ContentPost.objects.all().delete()
            self.stdout.write(self.style.WARNING('Cleared all existing content posts.'))

        wb = openpyxl.load_workbook(filepath)

        # ── 1. Master Idea Bank ────────────────────────────────────────────
        ws = wb['📋 Master Idea Bank']
        posts_created = 0
        # Header row is row 5 (0-indexed: row 4)
        # Cols: #, Platform, Content Pillar, Tone, Hook, Full Copy, Visual Direction, Hashtags, Status, Priority
        for row in ws.iter_rows(min_row=6, values_only=True):
            if not row[0]:
                continue
            platform_raw = str(row[1]).strip() if row[1] else ''
            platform = PLATFORM_MAP.get(platform_raw, 'twitter')
            status_raw = str(row[8]).strip() if row[8] else 'To Do'
            priority_raw = str(row[9]).strip() if row[9] else 'Medium'

            ContentPost.objects.create(
                platform=platform,
                content_pillar=str(row[2] or '').strip(),
                tone=str(row[3] or '').strip(),
                hook=str(row[4] or '').strip(),
                full_copy=str(row[5] or '').strip(),
                visual_direction=str(row[6] or '').strip(),
                hashtags=str(row[7] or '').strip(),
                status=STATUS_MAP.get(status_raw, 'to_do'),
                priority=PRIORITY_MAP.get(priority_raw, 'medium'),
            )
            posts_created += 1

        self.stdout.write(self.style.SUCCESS(f'Imported {posts_created} content posts from Master Idea Bank.'))

        # ── 2. Content Calendar ────────────────────────────────────────────
        ws_cal = wb['📅 Content Calendar']
        cal_created = 0
        # Header: Week, Day, Date, Platform, Content Pillar, Hook, Copy, Visual Notes, Hashtags, Status, Posted?
        for row in ws_cal.iter_rows(min_row=5, values_only=True):
            if not row[3]:  # no platform = skip
                continue
            platform_raw = str(row[3]).strip()
            platform = PLATFORM_MAP.get(platform_raw, 'twitter')
            status_raw = str(row[9]).strip() if row[9] else 'To Do'

            ContentPost.objects.create(
                platform=platform,
                content_pillar=str(row[4] or '').strip(),
                hook=str(row[5] or '').strip(),
                full_copy=str(row[6] or '').strip(),
                visual_direction=str(row[7] or '').strip(),
                hashtags=str(row[8] or '').strip(),
                status=STATUS_MAP.get(status_raw, 'to_do'),
                week=str(row[0] or '').strip(),
                day=str(row[1] or '').strip(),
                scheduled_date=row[2] if row[2] else None,
            )
            cal_created += 1

        self.stdout.write(self.style.SUCCESS(
            f'Imported {cal_created} calendar entries.\n'
            f'Total: {posts_created + cal_created} content posts.'
        ))
